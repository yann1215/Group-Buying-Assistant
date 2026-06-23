# app/analysis/order_parser.py

from __future__ import annotations

import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config import CSV_OUTPUT_DIR, ensure_dirs


SHEET_INDEX = 1
HEADER_ROW = 3

ORDER_NO_COL = 1
NICKNAME_COL = 2

ORDER_NO_HEADER = "单号"
NICKNAME_HEADER = "昵称"
TOTAL_AMOUNT_HEADER = "总金额"
DEFAULT_PRODUCT_ANCHOR_HEADERS = (
    "发货状态",
    "卖家备注",
    "个人备注用户信息",
    "个人备注其他信息",
    "联系人电话",
    "收货人姓名",
    "收货人联系方式",
    "收货人地址",
)


class OrderParseError(RuntimeError):
    """订单表解析失败。"""


def parse_order_file(
    order_input: str | Path | dict[str, Any],
    output_dir: str | Path | None = None,
    product_anchor_header: str | None = None,
) -> str:
    """
    读取订单 Excel 的第 2 个 sheet，将订单商品数据整理为宽表 CSV。

    输入表规则：
        - 第 2 个 sheet
        - 第 3 行是表头
        - 第 1 列是“单号”
        - 第 2 列是“昵称”
        - 从第 3 行最右侧向左查找商品定位关键词
        - 找到关键词后，关键词所在列之后均为商品名称列
        - 商品数量只能是正整数或空
        - 单号必须是正整数，不能为空

    输出 CSV：
        单号, 昵称, 商品1, 商品2, 商品3, ...

    返回：
        处理后的 CSV 文件绝对路径。
    """
    input_path = _get_input_path(order_input)

    if not input_path.exists():
        raise FileNotFoundError(f"订单文件不存在：{input_path}")

    if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise OrderParseError(
            f"暂时只支持 .xlsx / .xlsm 文件。当前文件：{input_path.name}"
        )

    ensure_dirs()

    output_dir_path = Path(output_dir) if output_dir else CSV_OUTPUT_DIR
    output_dir_path.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(input_path, data_only=True)

    if len(wb.worksheets) <= SHEET_INDEX:
        raise OrderParseError(
            f"Excel 中没有第 2 个 sheet。当前 sheet 数量：{len(wb.worksheets)}"
        )

    ws = wb.worksheets[SHEET_INDEX]
    merged_value_map = _build_merged_value_map(ws)

    _validate_fixed_headers(ws, merged_value_map)

    # 定位订单总金额列
    total_amount_col = _find_required_header_col(
        ws=ws,
        header_name=TOTAL_AMOUNT_HEADER,
        merged_value_map=merged_value_map,
    )

    # 如果外部明确指定了定位表头，只查找指定表头；
    # 否则使用默认候选关键词。
    candidate_anchor_headers = (
        (product_anchor_header,)
        if product_anchor_header
        else DEFAULT_PRODUCT_ANCHOR_HEADERS
    )

    # 从表头行最右侧向左查找。
    # 找到任意候选关键词后，该关键词下一列即为商品起始列。
    product_anchor_col, matched_anchor_header = _find_product_anchor_col(
        ws=ws,
        header_names=candidate_anchor_headers,
        merged_value_map=merged_value_map,
    )

    product_names = _get_product_names(
        ws=ws,
        start_col=product_anchor_col + 1,
        merged_value_map=merged_value_map,
    )

    if not product_names:
        raise OrderParseError(
            f"已找到商品定位表头“{matched_anchor_header}”，"
            f"但没有在该列之后找到商品名称。"
        )

    product_headers = _make_unique_headers(product_names)

    # timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # # 考虑到前后的表格可能同名，并且在同一天进行处理，因此保留了h-m-s的后缀，用以保证区分文件
    # output_path = output_dir_path / f"{input_path.stem}_parsed_orders_{timestamp}.csv"
    output_path = output_dir_path / f"{input_path.stem}_parsed_orders.csv"

    fieldnames = ["单号", "昵称", "总金额"] + product_headers

    rows_written = 0

    with output_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for row_idx in range(HEADER_ROW + 1, ws.max_row + 1):
            if _is_empty_data_row(
                ws=ws,
                row=row_idx,
                last_col=product_anchor_col + len(product_names),
                merged_value_map=merged_value_map,
            ):
                continue

            order_no_raw = _get_cell_value(
                ws=ws,
                row=row_idx,
                col=ORDER_NO_COL,
                merged_value_map=merged_value_map,
            )
            nickname_raw = _get_cell_value(
                ws=ws,
                row=row_idx,
                col=NICKNAME_COL,
                merged_value_map=merged_value_map,
            )

            total_amount_raw = _get_cell_value(
                ws=ws,
                row=row_idx,
                col=total_amount_col,
                merged_value_map=merged_value_map,
            )

            order_no = _parse_positive_int_required(
                value=order_no_raw,
                field_name="单号",
                row_idx=row_idx,
            )

            row_data: dict[str, Any] = {
                "单号": order_no,
                "昵称": _to_text(nickname_raw),
                "总金额": _format_total_amount(total_amount_raw),
            }

            for offset, product_header in enumerate(product_headers):
                col_idx = product_anchor_col + 1 + offset

                quantity_raw = _get_cell_value(
                    ws=ws,
                    row=row_idx,
                    col=col_idx,
                    merged_value_map=merged_value_map,
                )

                quantity = _parse_quantity_optional(
                    value=quantity_raw,
                    row_idx=row_idx,
                    product_name=product_header,
                )

                row_data[product_header] = quantity

            writer.writerow(row_data)
            rows_written += 1

    if rows_written == 0:
        raise OrderParseError("解析完成，但没有提取到任何订单行。")

    return str(output_path.resolve())


def _get_input_path(order_input: str | Path | dict[str, Any]) -> Path:
    if isinstance(order_input, (str, Path)):
        return Path(order_input)

    if isinstance(order_input, dict):
        for key in ("file_path", "order_file", "path"):
            value = order_input.get(key)
            if value:
                return Path(value)

    raise ValueError(
        "order_input 必须是文件路径，或包含 file_path / order_file / path 的 dict。"
    )


def _build_merged_value_map(ws: Worksheet) -> dict[tuple[int, int], Any]:
    """
    让合并单元格区域内的每个位置都能读到左上角的值。
    """
    merged_value_map: dict[tuple[int, int], Any] = {}

    for merged_range in ws.merged_cells.ranges:
        top_left_value = ws.cell(
            row=merged_range.min_row,
            column=merged_range.min_col,
        ).value

        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_value_map[(row, col)] = top_left_value

    return merged_value_map


def _get_cell_value(
    ws: Worksheet,
    row: int,
    col: int,
    merged_value_map: dict[tuple[int, int], Any],
) -> Any:
    if (row, col) in merged_value_map:
        return merged_value_map[(row, col)]

    return ws.cell(row=row, column=col).value


def _validate_fixed_headers(
    ws: Worksheet,
    merged_value_map: dict[tuple[int, int], Any],
) -> None:
    """
    检查第 3 行第 1 列、第 2 列是否分别为“单号”“昵称”。
    """
    order_no_header = _get_cell_value(ws, HEADER_ROW, ORDER_NO_COL, merged_value_map)
    nickname_header = _get_cell_value(ws, HEADER_ROW, NICKNAME_COL, merged_value_map)

    if _normalize_header(order_no_header) != ORDER_NO_HEADER:
        raise OrderParseError(
            f"第 {HEADER_ROW} 行第 {ORDER_NO_COL} 列应为“{ORDER_NO_HEADER}”，"
            f"实际为“{_to_text(order_no_header)}”。"
        )

    if _normalize_header(nickname_header) != NICKNAME_HEADER:
        raise OrderParseError(
            f"第 {HEADER_ROW} 行第 {NICKNAME_COL} 列应为“{NICKNAME_HEADER}”，"
            f"实际为“{_to_text(nickname_header)}”。"
        )


def _find_required_header_col(
    ws: Worksheet,
    header_name: str,
    merged_value_map: dict[tuple[int, int], Any],
) -> int:
    """
    在第 3 行中查找指定表头。

    用于查找“总金额”等必须存在、但列位置不固定的字段。
    """
    target = _normalize_header(header_name)

    for col in range(1, ws.max_column + 1):
        value = _get_cell_value(
            ws=ws,
            row=HEADER_ROW,
            col=col,
            merged_value_map=merged_value_map,
        )

        if _normalize_header(value) == target:
            return col

    raise OrderParseError(
        f"第 {HEADER_ROW} 行没有找到表头“{header_name}”。"
    )


def _find_product_anchor_col(
    ws: Worksheet,
    header_names: tuple[str, ...],
    merged_value_map: dict[tuple[int, int], Any],
) -> tuple[int, str]:
    """
    从第 3 行最右侧向左查找商品定位表头。

    找到以下任意候选表头后立即停止：
    发货状态、卖家备注、个人备注用户信息、个人备注其他信息、
    联系人电话、收货人姓名、收货人联系方式、收货人地址。

    返回：
        定位表头所在列号、匹配到的标准表头名称。
    """
    normalized_headers: dict[str, str] = {}

    for header_name in header_names:
        normalized_name = _normalize_header(header_name)
        if normalized_name:
            normalized_headers[normalized_name] = header_name

    # 从最后一列向第一列查找
    for col in range(ws.max_column, 0, -1):
        value = _get_cell_value(
            ws=ws,
            row=HEADER_ROW,
            col=col,
            merged_value_map=merged_value_map,
        )

        normalized_value = _normalize_header(value)

        if normalized_value in normalized_headers:
            return col, normalized_headers[normalized_value]

    expected_headers = "、".join(
        f"“{header_name}”" for header_name in header_names
    )

    raise OrderParseError(
        f"第 {HEADER_ROW} 行没有找到任何商品定位表头。"
        f"可识别的表头包括：{expected_headers}"
    )


def _get_product_names(
    ws: Worksheet,
    start_col: int,
    merged_value_map: dict[tuple[int, int], Any],
) -> list[str]:
    """
    从商品起始列开始，读取所有非空商品名称。
    """
    product_names: list[str] = []

    for col in range(start_col, ws.max_column + 1):
        product_name = _get_cell_value(ws, HEADER_ROW, col, merged_value_map)

        if _is_blank(product_name):
            continue

        product_names.append(_to_text(product_name))

    return product_names


def _make_unique_headers(headers: list[str]) -> list[str]:
    """
    CSV 不适合出现重复表头。

    如果商品名称重复：
        雪梅蜂
        雪梅蜂

    会改成：
        雪梅蜂
        雪梅蜂__2
    """
    seen: dict[str, int] = {}
    result: list[str] = []

    for header in headers:
        base = header.strip()

        if base not in seen:
            seen[base] = 1
            result.append(base)
        else:
            seen[base] += 1
            result.append(f"{base}__{seen[base]}")

    return result


def _is_empty_data_row(
    ws: Worksheet,
    row: int,
    last_col: int,
    merged_value_map: dict[tuple[int, int], Any],
) -> bool:
    """
    判断一行是否是完全空行。
    """
    for col in range(1, last_col + 1):
        value = _get_cell_value(ws, row, col, merged_value_map)
        if not _is_blank(value):
            return False

    return True


def _parse_positive_int_required(
    value: Any,
    field_name: str,
    row_idx: int,
) -> int:
    """
    必填正整数字段。

    用于“单号”。
    """
    parsed = _parse_positive_int(value)

    if parsed is None:
        raise OrderParseError(
            f"第 {row_idx} 行“{field_name}”必须是正整数，不能为空。"
            f"实际值：{value!r}"
        )

    return parsed


def _parse_quantity_optional(
    value: Any,
    row_idx: int,
    product_name: str,
) -> int | str:
    """
    商品数量允许为空。

    非空时必须是正整数。
    """
    if _is_blank(value):
        return ""

    parsed = _parse_positive_int(value)

    if parsed is None:
        raise OrderParseError(
            f"第 {row_idx} 行商品“{product_name}”的数量必须是正整数或空。"
            f"实际值：{value!r}"
        )

    return parsed


def _parse_positive_int(value: Any) -> int | None:
    """
    解析正整数。

    接受：
        1
        1.0
        "1"

    不接受：
        0
        -1
        1.5
        "1件"
        "一"
        ""
    """
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, int):
        return value if value > 0 else None

    if isinstance(value, float):
        if value.is_integer() and value > 0:
            return int(value)
        return None

    s = str(value).strip()

    if not re.fullmatch(r"\d+", s):
        return None

    parsed = int(s)

    if parsed <= 0:
        return None

    return parsed


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""

    return re.sub(r"\s+", "", str(value).strip())


def _is_blank(value: Any) -> bool:
    if value is None:
        return True

    if isinstance(value, str) and value.strip() == "":
        return True

    return False


def _to_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _format_total_amount(value: Any) -> str:
    """
    格式化订单文件中的总金额。

    规则：
    - 空值保持为空
    - 数字统一输出为两位小数
    - 非数字内容原样转成文本，避免因为格式异常导致整个订单解析失败
    """
    if _is_blank(value):
        return ""

    if isinstance(value, bool):
        return _to_text(value)

    if isinstance(value, (int, float)):
        return f"{value:.2f}"

    text = str(value).strip()

    try:
        number = float(text.replace(",", ""))
    except ValueError:
        return text

    return f"{number:.2f}"


if __name__ == "__main__":

    output_path = parse_order_file(
        order_input=r"D:\2_PycharmTestData\test\miao1.xlsx",
        output_dir=r"D:\2_PycharmTestData\test2",
    )

    print("订单解析完成：")
    print(output_path)
